#!/usr/bin/python
from .logAndPrint import logAndPrint
# @input 
# inputf: input file
# @return: A list of title
def queryOrder(inputf):
    with open(inputf) as f:
        titles = f.read().splitlines()

    titleList=[]
    for title in titles:
        if title.startswith(">"):
            title = title.replace(">", "", 1)
            titleList.append(title)
    
    return titleList

# @input 
# inputf: input file; cf: COI file; outputf: outfut file; logf: log file
def vsearchShower(inputf, cf, outputf, logf):
    titleList = queryOrder(inputf)

    with open(cf) as f:
        cresults = f.read().splitlines()
    dic_crts = {}
    for crt in cresults:
        list_crt = crt.split("\t")
        source_title = list_crt[0]
        target_title = list_crt[1]
        score = list_crt[2]+'%'
        if dic_crts.get(source_title, "null") == "null":
            dic_crts[source_title] = [(target_title, score),]
        else:
            dic_crts[source_title] += [(target_title, score),]

    with open(outputf, 'a+') as f:
        i=0
        for title in titleList:
            i+=1
            if not title in dic_crts:
                f.write((';').join([str(i), title, 'null', 'null', 'null', 'null', 'null', 'null', 'null', '0%', '\n']))
            else:
                target_matches = dic_crts[title]
                for target_match in target_matches:
                    target_title = target_match[0]
                    score = target_match[1]
                    list_levels = target_title.split('__')
                    f.write((';').join([str(i), title, list_levels[0], list_levels[1], list_levels[2], list_levels[3], list_levels[4], list_levels[5], list_levels[6], score, '\n']))
            #logAndPrint(logf, "...\n[ " + title + "] is successfully finished classifing!")

# @input
# inputf: input file; cf: COI file; outputf: outfut file; logf: log file
def blastnShower(inputf, cf, outputf, logf):
    with open(cf) as f:
        citems = f.read().splitlines()

    with open(cf, 'w+') as f:
        for citem in citems:
            if not citem.startswith("#"):
                f.write(citem+'\n')
    vsearchShower(inputf, cf, outputf, logf)


# @input 
# inputf: input file; tf: tsv file; outputf: outfut file; logf: log file
def gappaShower(inputf, tf, outputf, logf):
    titleList = queryOrder(inputf)

    with open(tf) as f:
        items = f.read().splitlines()
    dic_its={}
    for i in range(len(items)-1,0,-1):
        item = items[i]
        title = item.split('\t')[0]
        sem_count = item.count(';')
        if item.split('\t')[-1] != "DISTANT":
            if dic_its.get(title,'null_pot') == 'null_pot':
                dic_its[title] = {item: sem_count}
            elif sem_count in dic_its[title].values():
                dic_its[title][item] = sem_count

    with open(outputf,'a+') as f:
        i = 0
        for title in titleList:
            i += 1
            if not title in dic_its:
                f.write((';').join([str(i), title, 'null;'*7, '0\n']))
            else:
                for j in dic_its[title]:
                    levels = (j.split('\t')[5]).split(';')
                    score = j.split('\t')[3]
                    sem_count = int(dic_its[title][j])
                    kingdom = levels[0]
                    phylum = clas = order = family = genus = species  = 'null'
                    if sem_count > 0:
                        phylum = levels[1]
                    if sem_count > 1:
                        clas = levels[2]
                    if sem_count > 2:
                        order = levels[3]
                    if sem_count > 3:
                        family = levels[4]
                    if sem_count > 4:
                        genus = levels[5]
                    if sem_count > 5:
                        species = levels[6]
                    f.write((';').join([str(i), title, kingdom, phylum, clas, order, family, genus, species, '', score+'\n']))

                    #logAndPrint(logf, "...\n[ " + title + "] is successfully finished classifing!")

def uniqShower(inputf):
    with open(inputf) as f:
        items = f.read().splitlines()

    dic_items = {}
    for item in items:
        if item.startswith("NO."):
            dic_items[0] = item
        else:
            its = item.split(";")
            nb = its[0]
            sc = float(its[-1]) if its[-1] else 0
            if nb not in dic_items:
                dic_items[nb] = item
            elif sc > float(dic_items[nb].split(";")[-1]):
                dic_items[nb] = item

    with open(inputf, 'w+') as f:
        for i in dic_items.values():
            f.write(i+'\n')

def excelShower(outputf):
    if outputf.endswith(".xlsx"):
        from openpyxl import Workbook
        from openpyxl.utils import get_column_letter
        from openpyxl.styles import Font
        with open(outputf) as f:
            items = f.read().splitlines()
        wb = Workbook()
        ws = wb.active
        ws.title = 'result'
        for i in range(len(items)):
            its = items[i].split(";")
            for j in range(len(its)):
                letter = get_column_letter(j+1)
                if i == 0:
                    ws[letter+str(i+1)].font = Font(name='等线',size=11,bold=True)
                else:
                    ws[letter+str(i+1)].font = Font(name='等线',size=11)
                ws[letter+str(i+1)] = str(its[j])
        wb.save(outputf)
        wb.close()
